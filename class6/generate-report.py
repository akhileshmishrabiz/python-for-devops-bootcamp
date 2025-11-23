from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
import boto3

rds = boto3.client("rds")
cloudwatch = boto3.client("cloudwatch")


def save_to_excel(data, headers, excel_file):
    """Save data to an Excel file."""
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        ws.delete_rows(1, ws.max_row)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active

    # Write headers
    header_font = Font(bold=True, size=12)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font

    # Write data
    for row, row_data in enumerate(data, start=2):
        for col, cell_data in enumerate(row_data, start=1):
            ws.cell(row=row, column=col, value=str(cell_data))
            column_letter = get_column_letter(col)
            column_width = max(len(str(cell_data)), len(headers[col - 1]))
            ws.column_dimensions[column_letter].width = column_width + 2

    # Add table formatting
    table_range = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    table_name = f"Table_{timestamp}"

    tab = Table(displayName=table_name, ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    wb.save(excel_file)
    print(f"Saved report to {excel_file}")


def get_db_freestorage(db_identifier):
    """Get free storage space for RDS instance in GB."""
    response = cloudwatch.get_metric_data(
        MetricDataQueries=[
            {
                "Id": "fetching_FreeStorageSpace",
                "MetricStat": {
                    "Metric": {
                        "Namespace": "AWS/RDS",
                        "MetricName": "FreeStorageSpace",
                        "Dimensions": [
                            {"Name": "DBInstanceIdentifier", "Value": db_identifier}
                        ],
                    },
                    "Period": 86400 * 7,
                    "Stat": "Minimum",
                },
            }
        ],
        StartTime=(datetime.now() - timedelta(days=7)).timestamp(),
        EndTime=datetime.now().timestamp(),
        ScanBy="TimestampDescending",
    )

    try:
        return round(response["MetricDataResults"][0]["Values"][0] / 1024**3, 2)
    except (IndexError, KeyError):
        print(f"Error fetching free storage for {db_identifier}")
        return 0.0


def get_rds_instances():
    """Get all non-serverless, non-cluster RDS instances."""
    print("Fetching RDS instances...")
    db_instances = rds.describe_db_instances()["DBInstances"]

    filtered_instances = []
    for db in db_instances:
        if db.get("DBInstanceClass") != "db.serverless" and not db.get(
            "DBClusterIdentifier"
        ):
            filtered_instances.append(db)

    print(f"Found {len(filtered_instances)} RDS instances")
    return filtered_instances


def format_data(db_instances):
    """Format RDS data for Excel output."""
    data = []
    for db in db_instances:
        print(f"Processing {db['DBInstanceIdentifier']}...")
        row = (
            db["DBInstanceIdentifier"],
            db["DBInstanceClass"],
            db["Engine"],
            db["EngineVersion"],
            db["AllocatedStorage"],
            get_db_freestorage(db["DBInstanceIdentifier"]),
        )
        data.append(row)
    return data


def main():
    """Main function."""
    output_file = "rds_storage_report.xlsx"

    instances = get_rds_instances()
    data = format_data(instances)

    headers = [
        "DBInstanceIdentifier",
        "DBInstanceClass",
        "DBEngine",
        "Version",
        "AllocatedStorage",
        "FreeStorageSpace",
    ]

    save_to_excel(data, headers, output_file)
    print("Done!")


if __name__ == "__main__":
    main()
